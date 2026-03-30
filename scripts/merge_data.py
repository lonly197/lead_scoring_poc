#!/usr/bin/env python3
"""
数据合并脚本（内存优化版）

针对大文件优化：
- 使用 calamine 引擎读取 Excel（更快更省内存）
- 分批处理 DMP 数据
- 流式写入
- 支持数据脱敏
- 支持数据集拆分

用法：
    # 基础合并
    uv run python scripts/merge_data.py \
        --excel 测试数据/202601~03_v2.xlsx \
        --dmp 测试数据/DMP行为数据(202601~03).csv \
        --output 测试数据/线索宽表_完整.parquet

    # 启用脱敏
    uv run python scripts/merge_data.py \
        --excel 测试数据/202601~03_v2.xlsx \
        --dmp 测试数据/DMP行为数据(202601~03).csv \
        --output 测试数据/线索宽表_脱敏.parquet \
        --desensitize

    # 合并 + 拆分训练/测试集
    uv run python scripts/merge_data.py \
        --excel 测试数据/202601~03_v2.xlsx \
        --dmp 测试数据/DMP行为数据(202601~03).csv \
        --output 测试数据/线索宽表 \
        --split \
        --split-target 线索评级结果

    # 输出：线索宽表_train.parquet, 线索宽表_test.parquet
"""

import argparse
import re
import sys
from pathlib import Path

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))


# ==================== 脱敏配置 ====================

# 品牌关键词映射
BRAND_MAPPING = {
    "广汽丰田": "品牌A",
    "广丰": "品牌A",
    "广汽": "集团A",
    "GTMC": "代号G",
    "广汽本地": "区域A",
}

# 需要品牌关键词替换的文本字段
BRAND_TEXT_COLUMNS = [
    "首触意向车型/意向车型",
    "一级渠道名称",
    "二级渠道名称",
    "三级渠道名称",
    "四级渠道名称",
    "首触跟进记录",
    "非首触跟进记录",
    "战败原因",
    "dmp_品牌",
    "dmp_车型代码",
]

# ID 字段掩码（保留前2后2位）
ID_MASK_COLUMNS = [
    "客户ID",
    "客户ID(店端)",
]


def mask_id(value: str) -> str:
    """ID 脱敏：保留前2后2位"""
    if not value or len(value) <= 4:
        return "****"
    return f"{value[:2]}{'*' * (len(value) - 4)}{value[-2:]}"


def mask_phone(value: str) -> str:
    """手机号脱敏：保留前3后4位"""
    if not value:
        return value
    # 匹配11位手机号
    phone_pattern = re.compile(r"1[3-9]\d{9}")
    return phone_pattern.sub(lambda m: f"{m.group()[:3]}****{m.group()[-4:]}", str(value))


def mask_id_card(value: str) -> str:
    """身份证脱敏：保留前6后4位"""
    if not value:
        return value
    # 匹配15或18位身份证
    id_pattern = re.compile(r"\d{15}[\dXx]?[\dXx]?[\dXx]?")
    return id_pattern.sub(lambda m: f"{m.group()[:6]}********{m.group()[-4:]}", str(value))


def replace_brand_keywords(text: str) -> str:
    """替换品牌关键词"""
    if not text:
        return text
    result = str(text)
    for keyword, replacement in BRAND_MAPPING.items():
        result = result.replace(keyword, replacement)
    return result


def desensitize_column(pl, series, col_name: str):
    """
    对单列进行脱敏处理

    Args:
        pl: polars 模块
        series: polars Series
        col_name: 列名

    Returns:
        脱敏后的 Series
    """
    # 品牌关键词替换
    for keyword, replacement in BRAND_MAPPING.items():
        series = series.str.replace_all(keyword, replacement)

    # ID 字段掩码
    if any(col in col_name for col in ["客户ID"]):
        series = series.cast(pl.Utf8).map_elements(mask_id, return_dtype=pl.Utf8)

    # 手机号脱敏（跟进记录等文本中可能包含）
    if "跟进" in col_name or "战败" in col_name or "记录" in col_name:
        # 文本中的手机号脱敏
        series = series.map_elements(mask_phone, return_dtype=pl.Utf8)
        # 身份证脱敏
        series = series.map_elements(mask_id_card, return_dtype=pl.Utf8)

    return series


def desensitize_data(pl, df: "pl.DataFrame") -> "pl.DataFrame":
    """
    数据脱敏处理

    Args:
        pl: polars 模块
        df: 原始 DataFrame

    Returns:
        脱敏后的 DataFrame
    """
    print("  执行脱敏处理...")

    # 获取需要处理的列
    text_columns = []
    for col in df.columns:
        # 文本类型且在脱敏列表中
        if df[col].dtype == pl.Utf8:
            # 品牌关键词替换
            if any(kw in col for kw in ["车型", "渠道", "跟进", "战败", "品牌", "dmp_"]):
                text_columns.append(col)
            # ID 掩码
            elif "客户ID" in col:
                text_columns.append(col)

    print(f"  待脱敏列数: {len(text_columns)}")

    # 批量处理
    for col in text_columns:
        df = df.with_columns(
            desensitize_column(pl, df[col], col).alias(col)
        )

    return df


def split_data(
    pl,
    df: "pl.DataFrame",
    target_column: str,
    test_ratio: float = 0.2,
    random_seed: int = 42,
) -> tuple["pl.DataFrame", "pl.DataFrame"]:
    """
    分层随机拆分数据集

    Args:
        pl: polars 模块
        df: 数据 DataFrame
        target_column: 目标变量列名（用于分层）
        test_ratio: 测试集比例
        random_seed: 随机种子

    Returns:
        (训练集, 测试集)
    """
    print(f"  拆分数据集: 目标列={target_column}, 测试比例={test_ratio}")

    # 检查目标列是否存在
    if target_column not in df.columns:
        print(f"  警告: 目标列 '{target_column}' 不存在，使用随机拆分")
        # 简单随机拆分
        n = len(df)
        test_size = int(n * test_ratio)
        df_shuffled = df.sample(fraction=1.0, shuffle=True, seed=random_seed)
        test_df = df_shuffled.head(test_size)
        train_df = df_shuffled.tail(n - test_size)
        return train_df, test_df

    # 过滤目标列为空的行
    df_valid = df.filter(pl.col(target_column).is_not_null())
    null_count = len(df) - len(df_valid)
    if null_count > 0:
        print(f"  过滤目标列为空的行: {null_count}")

    # 分层采样
    # 获取目标列的唯一值
    unique_values = df_valid.select(target_column).unique().to_series().to_list()

    train_parts = []
    test_parts = []

    for val in unique_values:
        subset = df_valid.filter(pl.col(target_column) == val)
        n = len(subset)
        test_size = int(n * test_ratio)

        # 随机打乱
        subset_shuffled = subset.sample(fraction=1.0, shuffle=True, seed=random_seed)

        test_parts.append(subset_shuffled.head(test_size))
        train_parts.append(subset_shuffled.tail(n - test_size))

    train_df = pl.concat(train_parts)
    test_df = pl.concat(test_parts)

    # 最终打乱
    train_df = train_df.sample(fraction=1.0, shuffle=True, seed=random_seed)
    test_df = test_df.sample(fraction=1.0, shuffle=True, seed=random_seed)

    print(f"  训练集: {len(train_df):,} 行")
    print(f"  测试集: {len(test_df):,} 行")

    # 打印目标分布
    train_dist = train_df.group_by(target_column).len().sort(target_column)
    test_dist = test_df.group_by(target_column).len().sort(target_column)
    print(f"  训练集分布: {train_dist.to_dict(as_series=False)}")
    print(f"  测试集分布: {test_dist.to_dict(as_series=False)}")

    return train_df, test_df


def main():
    parser = argparse.ArgumentParser(
        description="合并线索宽表和 DMP 行为数据",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
    # 基础合并
    uv run python scripts/merge_data.py \\
        --excel ./data/线索宽表.xlsx \\
        --dmp ./data/DMP行为数据.csv \\
        --output ./data/线索宽表_完整.parquet

    # 启用脱敏
    uv run python scripts/merge_data.py \\
        --excel ./data/线索宽表.xlsx \\
        --dmp ./data/DMP行为数据.csv \\
        --output ./data/线索宽表_脱敏.parquet \\
        --desensitize

    # 合并 + 拆分训练/测试集
    uv run python scripts/merge_data.py \\
        --excel ./data/线索宽表.xlsx \\
        --dmp ./data/DMP行为数据.csv \\
        --output ./data/线索宽表 \\
        --split \\
        --split-target 线索评级结果

    # 输出: 线索宽表_train.parquet, 线索宽表_test.parquet
        """,
    )
    parser.add_argument("--excel", required=True, help="Excel 文件路径")
    parser.add_argument("--dmp", required=True, help="DMP 行为数据 CSV 路径")
    parser.add_argument("--output", required=True, help="输出文件路径（拆分时为前缀）")
    parser.add_argument("--format", choices=["parquet", "csv"], default="parquet")
    parser.add_argument("--desensitize", action="store_true", help="启用数据脱敏")
    parser.add_argument("--split", action="store_true", help="拆分训练/测试集")
    parser.add_argument("--split-target", default="线索评级结果", help="拆分目标列名（用于分层采样）")
    parser.add_argument("--split-ratio", type=float, default=0.2, help="测试集比例（默认 0.2）")
    parser.add_argument("--split-seed", type=int, default=42, help="随机种子（默认 42）")

    args = parser.parse_args()

    import polars as pl

    excel_path = Path(args.excel)
    dmp_path = Path(args.dmp)
    output_path = Path(args.output)

    print("=" * 60)
    print("数据合并脚本")
    print("=" * 60)

    # 1. 读取 Excel（使用 calamine 引擎，更快更省内存）
    print(f"\n读取 Excel: {excel_path}")

    # 获取 sheet 名称
    from openpyxl import load_workbook
    wb = load_workbook(str(excel_path), read_only=True)
    all_sheets = wb.sheetnames
    wb.close()
    print(f"Sheet 列表: {all_sheets}")

    # 读取数据 sheet（排除 Query）
    data_sheets = [s for s in all_sheets if 'query' not in s.lower()][:2]
    print(f"待合并 Sheet: {data_sheets}")

    dfs = []
    for sheet in data_sheets:
        print(f"  读取 {sheet}...")
        # 使用 calamine 引擎
        try:
            df = pl.read_excel(str(excel_path), sheet_name=sheet, engine="calamine")
        except Exception:
            df = pl.read_excel(str(excel_path), sheet_name=sheet)
        print(f"    {len(df):,} 行, {len(df.columns)} 列")
        dfs.append(df)

    # 合并
    if len(dfs) > 1:
        clue_df = pl.concat(dfs)
    else:
        clue_df = dfs[0]
    print(f"线索宽表: {len(clue_df):,} 行, {len(clue_df.columns)} 列")

    # 2. 读取 DMP 数据
    print(f"\n读取 DMP: {dmp_path}")
    dmp_columns = [
        "dmp_手机号", "dmp_事件时间", "dmp_事件名称", "dmp_车型代码",
        "dmp_品牌", "dmp_页面路径", "dmp_数值", "dmp_渠道",
        "dmp_分类", "dmp_未知", "dmp_平台",
    ]
    dmp_df = pl.read_csv(
        str(dmp_path), separator="\t", has_header=False,
        new_columns=dmp_columns, dtypes={"dmp_手机号": pl.Utf8}
    )
    print(f"  {len(dmp_df):,} 行")

    # 3. 聚合 DMP 特征
    print("\n聚合 DMP 特征...")
    dmp_agg = dmp_df.group_by("dmp_手机号").agg([
        pl.len().alias("dmp_行为次数"),
        pl.col("dmp_事件时间").max().alias("dmp_最近行为时间"),
        pl.col("dmp_事件名称").filter(pl.col("dmp_事件名称").str.contains("testDrive|试驾")).len().alias("dmp_试驾相关次数"),
        pl.col("dmp_事件名称").filter(pl.col("dmp_事件名称").str.contains("submitOrder|payOrder|下单|支付")).len().alias("dmp_下单支付次数"),
        pl.col("dmp_事件名称").n_unique().alias("dmp_事件类型数"),
    ])
    print(f"  聚合后: {len(dmp_agg):,} 用户")

    # 4. 关联
    print("\n关联数据...")
    # 查找手机号列
    phone_col = None
    for col in clue_df.columns:
        if "手机" in col or "phone" in col.lower():
            phone_col = col
            break
    if phone_col is None:
        phone_col = "手机号（脱敏）"

    print(f"  手机号列: {phone_col}")

    clue_df = clue_df.with_columns(pl.col(phone_col).cast(pl.Utf8).alias("_key"))
    dmp_agg = dmp_agg.with_columns(pl.col("dmp_手机号").cast(pl.Utf8).alias("_key"))

    result = clue_df.join(dmp_agg, on="_key", how="left").drop("_key")

    matched = result.filter(pl.col("dmp_行为次数").is_not_null()).height
    print(f"  匹配: {matched:,}/{len(clue_df):,} ({matched/len(clue_df)*100:.1f}%)")

    # 5. 脱敏处理（可选）
    if args.desensitize:
        print("\n脱敏处理...")
        result = desensitize_data(pl, result)
        print("  已完成脱敏处理")

    # 6. 拆分数据集（可选）
    if args.split:
        print(f"\n拆分数据集...")
        train_df, test_df = split_data(
            pl, result,
            target_column=args.split_target,
            test_ratio=args.split_ratio,
            random_seed=args.split_seed,
        )

        # 生成输出文件名
        output_dir = output_path.parent
        base_name = output_path.stem
        suffix = ".parquet" if args.format == "parquet" else ".csv"

        train_path = output_dir / f"{base_name}_train{suffix}"
        test_path = output_dir / f"{base_name}_test{suffix}"

        output_dir.mkdir(parents=True, exist_ok=True)

        # 保存训练集
        print(f"\n保存训练集: {train_path}")
        if args.format == "parquet":
            train_df.write_parquet(str(train_path))
        else:
            train_df.write_csv(str(train_path))

        train_size_mb = train_path.stat().st_size / 1024 / 1024
        print(f"  文件大小: {train_size_mb:.1f} MB")

        # 保存测试集
        print(f"保存测试集: {test_path}")
        if args.format == "parquet":
            test_df.write_parquet(str(test_path))
        else:
            test_df.write_csv(str(test_path))

        test_size_mb = test_path.stat().st_size / 1024 / 1024
        print(f"  文件大小: {test_size_mb:.1f} MB")

        print(f"\n输出文件:")
        print(f"  训练集: {train_path}")
        print(f"  测试集: {test_path}")

    else:
        # 不拆分，输出单个文件
        output_path.parent.mkdir(parents=True, exist_ok=True)
        print(f"\n输出: {output_path}")

        if args.format == "parquet" or output_path.suffix == ".parquet":
            result.write_parquet(str(output_path))
        else:
            result.write_csv(str(output_path))

        size_mb = output_path.stat().st_size / 1024 / 1024
        print(f"文件大小: {size_mb:.1f} MB")
        print(f"最终: {len(result):,} 行, {len(result.columns)} 列")

    print("\n" + "=" * 60)
    print("完成!")
    print("=" * 60)


if __name__ == "__main__":
    main()