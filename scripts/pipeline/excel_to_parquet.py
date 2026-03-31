#!/usr/bin/env python3
"""
Excel 转 Parquet 脚本（流式处理）

使用 openpyxl read_only 模式流式读取 Excel，
分批写入 Parquet，严格控制内存。

用法：
    uv run python scripts/pipeline/excel_to_parquet.py \
        --input ./data/large.xlsx \
        --output ./data/output.parquet \
        --batch-size 50000
"""

from __future__ import annotations

import argparse
import os
import sys
import time
from pathlib import Path
from typing import Iterator, List, Optional

# 强制禁用输出缓冲（必须在最开始设置）
os.environ['PYTHONUNBUFFERED'] = '1'

import polars as pl
import pyarrow.parquet as pq
from openpyxl import load_workbook


def log(msg: str = ""):
    """带时间戳的日志输出（强制刷新）"""
    timestamp = time.strftime("%H:%M:%S")
    print(f"[{timestamp}] {msg}", flush=True)


def iter_excel_rows(
    excel_path: Path,
    sheet_name: Optional[str] = None,
    header: bool = True,
) -> Iterator[List]:
    """
    流式读取 Excel 行（内存友好）
    """
    wb = load_workbook(str(excel_path), read_only=True, data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if header and i == 0:
            continue
        yield list(row)

    wb.close()


def get_excel_sheets(excel_path: Path) -> List[str]:
    """获取 Excel 所有 Sheet 名称"""
    wb = load_workbook(str(excel_path), read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def get_excel_header(excel_path: Path, sheet_name: Optional[str] = None) -> List[str]:
    """获取 Excel 表头"""
    wb = load_workbook(str(excel_path), read_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    wb.close()

    header = []
    for i, col in enumerate(header_row):
        if col is None or str(col).strip() == "":
            header.append(f"column_{i+1}")
        else:
            header.append(str(col).strip())

    return header


def excel_to_parquet_streaming(
    excel_path: Path,
    output_path: Path,
    sheet_names: Optional[List[str]] = None,
    batch_size: int = 50000,
    exclude_sheets: List[str] = None,
) -> Path:
    """流式转换 Excel 到 Parquet"""
    exclude_sheets = exclude_sheets or []
    start_time = time.time()

    log("=" * 50)
    log("Excel 流式转 Parquet")
    log("=" * 50)
    log(f"输入文件: {excel_path}")
    log(f"输出文件: {output_path}")
    log(f"批处理大小: {batch_size:,} 行/批")
    log("")

    # 获取 Sheet 列表
    all_sheets = get_excel_sheets(excel_path)
    log(f"Sheet 列表: {all_sheets}")

    # 自动过滤
    if sheet_names is None:
        sheet_names = [s for s in all_sheets
                       if s.lower() not in [e.lower() for e in exclude_sheets]
                       and 'query' not in s.lower()]

    log(f"待处理 Sheet: {sheet_names}")
    log("")

    # 确保输出目录存在
    output_path.parent.mkdir(parents=True, exist_ok=True)

    total_rows = 0
    batch_count = 0
    writer = None
    current_schema = None

    for sheet_idx, sheet_name in enumerate(sheet_names):
        log(f"[Sheet {sheet_idx+1}/{len(sheet_names)}] {sheet_name}")
        log("-" * 40)

        # 获取表头
        header = get_excel_header(excel_path, sheet_name)
        log(f"  列数: {len(header)}")

        sheet_start = time.time()
        sheet_rows = 0
        batch_rows = []

        # 流式读取行
        for row in iter_excel_rows(excel_path, sheet_name, header=True):
            batch_rows.append(row)

            # 批处理
            if len(batch_rows) >= batch_size:
                # 转换为 Polars DataFrame
                df = pl.DataFrame(batch_rows, schema=header, orient="row")
                table = df.to_arrow()

                # 初始化或验证 schema
                if writer is None:
                    current_schema = table.schema
                    writer = pq.ParquetWriter(output_path, current_schema)
                elif table.schema != current_schema:
                    table = table.cast(current_schema)

                # 写入批次
                writer.write_batch(table)

                batch_count += 1
                sheet_rows += len(batch_rows)
                total_rows += len(batch_rows)

                # 进度输出
                elapsed = time.time() - start_time
                rate = total_rows / elapsed if elapsed > 0 else 0
                log(f"  批次 {batch_count}: {sheet_rows:,} 行 ({rate:.0f} 行/秒)")

                # 清理
                del df, table, batch_rows
                batch_rows = []

        # 处理剩余行
        if batch_rows:
            df = pl.DataFrame(batch_rows, schema=header, orient="row")
            table = df.to_arrow()

            if writer is None:
                current_schema = table.schema
                writer = pq.ParquetWriter(output_path, current_schema)
            elif table.schema != current_schema:
                table = table.cast(current_schema)

            writer.write_batch(table)

            sheet_rows += len(batch_rows)
            total_rows += len(batch_rows)
            batch_count += 1
            del df, table, batch_rows

        sheet_elapsed = time.time() - sheet_start
        log(f"  完成: {sheet_rows:,} 行, 耗时 {sheet_elapsed:.1f}秒")
        log("")

    # 关闭 writer
    if writer:
        writer.close()

    # 摘要
    total_elapsed = time.time() - start_time
    log("=" * 50)
    log("转换完成")
    log("=" * 50)
    log(f"输出文件: {output_path}")
    log(f"文件大小: {output_path.stat().st_size / 1024 / 1024:.1f} MB")
    log(f"总行数: {total_rows:,}")
    log(f"总耗时: {total_elapsed:.1f} 秒")
    log(f"平均速度: {total_rows / total_elapsed:.0f} 行/秒")
    log("=" * 50)

    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Excel 流式转 Parquet（内存友好）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument(
        "--input", "-i",
        required=True,
        help="Excel 输入文件路径"
    )
    parser.add_argument(
        "--output", "-o",
        required=True,
        help="Parquet 输出文件路径"
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=50000,
        help="批处理大小（默认: 50000）"
    )
    parser.add_argument(
        "--sheets",
        nargs="+",
        default=None,
        help="指定处理的 Sheet"
    )
    parser.add_argument(
        "--exclude-sheets",
        nargs="+",
        default=["Query", "Summary"],
        help="排除的 Sheet 名称"
    )

    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        log(f"❌ 输入文件不存在: {input_path}")
        return 1

    try:
        excel_to_parquet_streaming(
            excel_path=input_path,
            output_path=output_path,
            sheet_names=args.sheets,
            batch_size=args.batch_size,
            exclude_sheets=args.exclude_sheets,
        )
        return 0

    except Exception as e:
        log(f"❌ 转换失败: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())