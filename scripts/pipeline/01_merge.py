#!/usr/bin/env python3
"""
数据合并脚本

功能：
- 读取 Excel 多 Sheet 并合并
- 读取 DMP 行为数据
- 按手机号关联，聚合 DMP 特征
- 输出合并后的 parquet 文件

内存控制：
- 使用 DuckDB 进行流式处理
- 可配置内存限制和线程数
- 避免全量加载到 Python 内存

用法：
    uv run python scripts/pipeline/01_merge.py \\
        --excel ./data/线索宽表.xlsx \\
        --dmp ./data/DMP行为数据.csv \\
        --output ./data/merged.parquet \\
        --memory-limit 4GB \\
        --threads 4
"""

from __future__ import annotations

import argparse
import sys
import tempfile
from pathlib import Path
from typing import List, Optional

# 添加项目根目录
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from src.pipeline.utils import print_step, format_size
from src.pipeline.config import DMP_COLUMNS


# ==================== DuckDB 流式合并 ====================

def merge_with_duckdb(
    excel_path: Path,
    dmp_path: Path,
    output_path: Path,
    phone_column: str = None,
    memory_limit: str = "4GB",
    threads: int = 4,
    sheet_limit: int = 2,
    temp_dir: Optional[Path] = None,
) -> Path:
    """
    使用 DuckDB 流式合并数据

    优势：
    - 内存限制：DuckDB 自动管理内存，不会超过设定值
    - 流式输出：直接写入 Parquet，不经过 Python DataFrame
    - 资源控制：可配置线程数，避免占满 CPU

    Args:
        excel_path: Excel 文件路径
        dmp_path: DMP 数据文件路径
        output_path: 输出文件路径
        phone_column: 手机号列名（自动检测如果为空）
        memory_limit: DuckDB 内存限制（如 "4GB"）
        threads: 线程数限制
        sheet_limit: 最多读取的 Sheet 数量
        temp_dir: 临时文件目录

    Returns:
        输出文件路径
    """
    import duckdb

    print("=" * 60)
    print("DuckDB 流式数据合并")
    print("=" * 60)
    print(f"  内存限制: {memory_limit}")
    print(f"  线程数: {threads}")

    # 创建 DuckDB 连接（内存模式）
    con = duckdb.connect(":memory:")

    # 设置资源限制
    con.execute(f"SET memory_limit='{memory_limit}'")
    con.execute(f"SET threads={threads}")

    # 临时目录
    if temp_dir is None:
        temp_dir = Path(tempfile.mkdtemp(prefix="merge_"))
    temp_dir.mkdir(parents=True, exist_ok=True)

    try:
        # 1. 安装并加载 Excel 扩展
        print_step("加载 Excel 扩展", "running")
        try:
            con.execute("INSTALL excel")
            con.execute("LOAD excel")
            print_step("加载 Excel 扩展", "success")
        except Exception as e:
            print_step("加载 Excel 扩展", "error", str(e))
            raise RuntimeError("无法加载 DuckDB Excel 扩展，请使用 --mode polars")

        # 2. 获取 Excel Sheet 列表
        print_step("扫描 Excel", "running", str(excel_path))
        sheets_result = con.execute(f"""
            SELECT sheet_name FROM excel_sheets('{excel_path}')
        """).fetchall()
        all_sheets = [row[0] for row in sheets_result]
        print(f"  Sheet 列表: {all_sheets}")

        # 读取数据 sheet（排除 Query）
        data_sheets = [s for s in all_sheets if 'query' not in s.lower()][:sheet_limit]
        print(f"  待合并 Sheet: {data_sheets}")
        print_step("扫描 Excel", "success", f"{len(data_sheets)} 个 Sheet")

        # 3. 读取并合并 Excel Sheet
        print_step("读取 Excel 数据", "running")

        clue_views = []
        for i, sheet in enumerate(data_sheets):
            view_name = f"clue_sheet_{i}"
            con.execute(f"""
                CREATE VIEW {view_name} AS
                SELECT * FROM read_xlsx('{excel_path}', sheet='{sheet}')
            """)

            # 获取行数
            row_count = con.execute(f"SELECT COUNT(*) FROM {view_name}").fetchone()[0]
            print(f"    {sheet}: {row_count:,} 行")
            clue_views.append(view_name)

        # 合并所有 Sheet
        if len(clue_views) > 1:
            union_sql = " UNION ALL ".join([f"SELECT * FROM {v}" for v in clue_views])
            con.execute(f"CREATE VIEW clue AS {union_sql}")
        else:
            con.execute(f"CREATE VIEW clue AS SELECT * FROM {clue_views[0]}")

        total_rows = con.execute("SELECT COUNT(*) FROM clue").fetchone()[0]
        total_cols = len(con.execute("SELECT * FROM clue LIMIT 0").description)
        print_step("读取 Excel 数据", "success", f"{total_rows:,} 行, {total_cols} 列")

        # 4. 自动检测手机号列
        print_step("检测手机号列", "running")
        columns = [col[0] for col in con.execute("SELECT * FROM clue LIMIT 0").description]

        if phone_column is None:
            phone_patterns = ["手机", "phone"]
            for col in columns:
                col_lower = col.lower()
                for pattern in phone_patterns:
                    if pattern in col_lower:
                        phone_column = col
                        break
                if phone_column:
                    break

        if phone_column is None:
            phone_column = "手机号（脱敏）"
            print(f"  警告: 未找到手机号列，使用默认: {phone_column}")
        else:
            print(f"  手机号列: {phone_column}")
        print_step("检测手机号列", "success")

        # 5. 读取并聚合 DMP 数据
        print_step("读取 DMP 数据", "running", str(dmp_path))

        # DMP 列定义（TSV 无表头）
        dmp_col_defs = ", ".join([f'column{i+1} AS {col}' for i, col in enumerate(DMP_COLUMNS)])

        con.execute(f"""
            CREATE VIEW dmp AS
            SELECT {dmp_col_defs}
            FROM read_csv('{dmp_path}', delim='\t', header=false)
        """)

        dmp_rows = con.execute("SELECT COUNT(*) FROM dmp").fetchone()[0]
        print_step("读取 DMP 数据", "success", f"{dmp_rows:,} 行")

        # 6. 聚合 DMP 特征
        print_step("聚合 DMP 特征", "running")

        con.execute("""
            CREATE VIEW dmp_agg AS
            SELECT
                dmp_手机号,
                COUNT(*) as dmp_行为次数,
                MAX(dmp_事件时间) as dmp_最近行为时间,
                COUNT(*) FILTER (
                    WHERE dmp_事件名称 LIKE '%testDrive%'
                       OR dmp_事件名称 LIKE '%试驾%'
                ) as dmp_试驾相关次数,
                COUNT(*) FILTER (
                    WHERE dmp_事件名称 LIKE '%submitOrder%'
                       OR dmp_事件名称 LIKE '%payOrder%'
                       OR dmp_事件名称 LIKE '%下单%'
                       OR dmp_事件名称 LIKE '%支付%'
                ) as dmp_下单支付次数,
                COUNT(DISTINCT dmp_事件名称) as dmp_事件类型数
            FROM dmp
            GROUP BY dmp_手机号
        """)

        dmp_agg_rows = con.execute("SELECT COUNT(*) FROM dmp_agg").fetchone()[0]
        print_step("聚合 DMP 特征", "success", f"{dmp_agg_rows:,} 用户")

        # 7. Left Join 关联
        print_step("关联数据", "running")

        # 确保输出目录存在
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # 直接输出到 Parquet（流式，不经过 Python 内存）
        con.execute(f"""
            COPY (
                SELECT
                    c.*,
                    d.dmp_行为次数,
                    d.dmp_最近行为时间,
                    d.dmp_试驾相关次数,
                    d.dmp_下单支付次数,
                    d.dmp_事件类型数
                FROM clue c
                LEFT JOIN dmp_agg d ON c."{phone_column}" = d.dmp_手机号
            ) TO '{output_path}' (FORMAT PARQUET, COMPRESSION ZSTD)
        """)

        # 获取匹配统计
        stats = con.execute(f"""
            SELECT
                COUNT(*) as total,
                COUNT(dmp_行为次数) as matched
            FROM read_parquet('{output_path}')
        """).fetchone()

        total, matched = stats
        match_rate = matched / total * 100 if total > 0 else 0
        print_step("关联数据", "success", f"匹配 {matched:,}/{total:,} ({match_rate:.1f}%)")

        # 8. 打印摘要
        print("\n" + "=" * 60)
        print("合并完成")
        print("=" * 60)
        print(f"输出文件: {output_path}")
        print(f"文件大小: {format_size(output_path)}")
        print(f"数据量: {total:,} 行")
        print(f"DMP 匹配率: {match_rate:.1f}%")
        print(f"内存限制: {memory_limit}")
        print(f"使用线程: {threads}")
        print("=" * 60)

        return output_path

    finally:
        con.close()


# ==================== Polars 传统合并（备选方案）====================

def merge_with_polars(
    excel_path: Path,
    dmp_path: Path,
    output_path: Path,
    sheet_limit: int = 2,
) -> Path:
    """
    使用 Polars 合并数据（备选方案）

    当 DuckDB Excel 扩展不可用时使用此方法。

    Args:
        excel_path: Excel 文件路径
        dmp_path: DMP 数据文件路径
        output_path: 输出文件路径
        sheet_limit: 最多读取的 Sheet 数量

    Returns:
        输出文件路径
    """
    import polars as pl

    print("=" * 60)
    print("Polars 数据合并")
    print("=" * 60)

    # 1. 读取 Excel
    print_step("读取 Excel", "running", str(excel_path))

    from openpyxl import load_workbook
    wb = load_workbook(str(excel_path), read_only=True)
    all_sheets = wb.sheetnames
    wb.close()

    print(f"  Sheet 列表: {all_sheets}")

    # 读取数据 sheet（排除 Query）
    data_sheets = [s for s in all_sheets if 'query' not in s.lower()][:sheet_limit]
    print(f"  待合并 Sheet: {data_sheets}")

    dfs = []
    for sheet in data_sheets:
        print(f"    读取 {sheet}...")
        df = None
        last_error = None
        for engine in ["openpyxl", "fastexcel"]:
            try:
                df = pl.read_excel(str(excel_path), sheet_name=sheet, engine=engine)
                print(f"      使用 {engine} 引擎")
                break
            except FileNotFoundError:
                raise
            except PermissionError:
                raise
            except Exception as e:
                last_error = e
                continue
        if df is None:
            raise RuntimeError(f"无法读取 Sheet: {sheet}, 最后错误: {last_error}")
        print(f"      {len(df):,} 行, {len(df.columns)} 列")
        dfs.append(df)

    # 合并
    if len(dfs) > 1:
        clue_df = pl.concat(dfs)
    else:
        clue_df = dfs[0]

    # 释放中间数据
    dfs.clear()
    del dfs

    print_step("读取 Excel", "success", f"{len(clue_df):,} 行, {len(clue_df.columns)} 列")

    # 2. 读取 DMP 数据
    print_step("读取 DMP", "running", str(dmp_path))

    dmp_df = pl.read_csv(
        str(dmp_path),
        separator="\t",
        has_header=False,
        new_columns=DMP_COLUMNS,
        dtypes={"dmp_手机号": pl.Utf8}
    )

    print_step("读取 DMP", "success", f"{len(dmp_df):,} 行")

    # 3. 聚合 DMP 特征
    print_step("聚合 DMP 特征", "running")

    dmp_agg = dmp_df.group_by("dmp_手机号").agg([
        pl.len().alias("dmp_行为次数"),
        pl.col("dmp_事件时间").max().alias("dmp_最近行为时间"),
        pl.col("dmp_事件名称").filter(
            pl.col("dmp_事件名称").str.contains("testDrive|试驾")
        ).len().alias("dmp_试驾相关次数"),
        pl.col("dmp_事件名称").filter(
            pl.col("dmp_事件名称").str.contains("submitOrder|payOrder|下单|支付")
        ).len().alias("dmp_下单支付次数"),
        pl.col("dmp_事件名称").n_unique().alias("dmp_事件类型数"),
    ])

    print_step("聚合 DMP 特征", "success", f"{len(dmp_agg):,} 用户")

    # 释放原始 DMP 数据
    del dmp_df

    # 4. 关联数据
    print_step("关联数据", "running")

    # 自动检测手机号列
    phone_col = None
    phone_patterns = ["手机", "phone"]
    for col in clue_df.columns:
        col_lower = col.lower()
        for pattern in phone_patterns:
            if pattern in col_lower:
                phone_col = col
                break
        if phone_col:
            break

    if phone_col is None:
        phone_col = "手机号（脱敏）"
        print(f"  警告: 未找到手机号列，使用默认: {phone_col}")
    else:
        print(f"  手机号列: {phone_col}")

    clue_df = clue_df.with_columns(pl.col(phone_col).cast(pl.Utf8).alias("_key"))
    dmp_agg = dmp_agg.with_columns(pl.col("dmp_手机号").cast(pl.Utf8).alias("_key"))

    result = clue_df.join(dmp_agg, on="_key", how="left").drop("_key")

    matched = result.filter(pl.col("dmp_行为次数").is_not_null()).height
    match_rate = matched / len(clue_df) * 100

    print_step("关联数据", "success", f"匹配 {matched:,}/{len(clue_df):,} ({match_rate:.1f}%)")

    # 5. 保存结果
    print_step("保存结果", "running", str(output_path))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    result.write_parquet(output_path, compression="zstd")

    print_step("保存结果", "success", format_size(output_path))

    # 打印摘要
    print("\n" + "=" * 60)
    print("合并完成")
    print("=" * 60)
    print(f"输出文件: {output_path}")
    print(f"文件大小: {format_size(output_path)}")
    print(f"数据量: {len(result):,} 行, {len(result.columns)} 列")
    print(f"DMP 匹配率: {match_rate:.1f}%")
    print("=" * 60)

    return output_path


# ==================== 主入口 ====================

def main() -> int:
    parser = argparse.ArgumentParser(
        description="数据合并脚本 - 合并 Excel 线索宽表和 DMP 行为数据",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
    # DuckDB 流式处理（推荐，内存可控）
    uv run python scripts/pipeline/01_merge.py \\
        --excel ./data/线索宽表.xlsx \\
        --dmp ./data/DMP行为数据.csv \\
        --output ./data/merged.parquet

    # 自定义内存和线程限制
    uv run python scripts/pipeline/01_merge.py \\
        --excel ./data/线索宽表.xlsx \\
        --dmp ./data/DMP行为数据.csv \\
        --output ./data/merged.parquet \\
        --memory-limit 2GB \\
        --threads 2

    # Polars 模式（DuckDB Excel 扩展不可用时）
    uv run python scripts/pipeline/01_merge.py \\
        --excel ./data/线索宽表.xlsx \\
        --dmp ./data/DMP行为数据.csv \\
        --output ./data/merged.parquet \\
        --mode polars
        """,
    )

    parser.add_argument(
        "--excel", "-e",
        required=True,
        help="Excel 文件路径（线索宽表）"
    )
    parser.add_argument(
        "--dmp", "-d",
        required=True,
        help="DMP 行为数据文件路径（CSV/TSV）"
    )
    parser.add_argument(
        "--output", "-o",
        default="./data/merged.parquet",
        help="输出文件路径（默认: data/merged.parquet）"
    )
    parser.add_argument(
        "--mode", "-m",
        choices=["duckdb", "polars"],
        default="duckdb",
        help="处理模式：duckdb（推荐，内存可控）或 polars（备选）"
    )
    parser.add_argument(
        "--memory-limit",
        default="4GB",
        help="DuckDB 内存限制（默认: 4GB）"
    )
    parser.add_argument(
        "--threads",
        type=int,
        default=4,
        help="线程数限制（默认: 4）"
    )
    parser.add_argument(
        "--sheet-limit",
        type=int,
        default=2,
        help="最多读取的 Sheet 数量（默认: 2）"
    )
    parser.add_argument(
        "--phone-column",
        default=None,
        help="手机号列名（自动检测如果为空）"
    )

    args = parser.parse_args()

    excel_path = Path(args.excel)
    dmp_path = Path(args.dmp)
    output_path = Path(args.output)

    # 检查输入文件
    if not excel_path.exists():
        print(f"❌ Excel 文件不存在: {excel_path}")
        return 1
    if not dmp_path.exists():
        print(f"❌ DMP 文件不存在: {dmp_path}")
        return 1

    try:
        if args.mode == "duckdb":
            merge_with_duckdb(
                excel_path=excel_path,
                dmp_path=dmp_path,
                output_path=output_path,
                phone_column=args.phone_column,
                memory_limit=args.memory_limit,
                threads=args.threads,
                sheet_limit=args.sheet_limit,
            )
        else:
            merge_with_polars(
                excel_path=excel_path,
                dmp_path=dmp_path,
                output_path=output_path,
                sheet_limit=args.sheet_limit,
            )
        return 0

    except Exception as e:
        print(f"❌ 合并失败: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())